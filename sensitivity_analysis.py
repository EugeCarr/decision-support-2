from matplotlib import pyplot as plt
import openpyxl
import numpy as np
from datetime import datetime
from simulation_0_0 import simulate
from openpyxl.styles import Font
import agent as ag
from regulator import Regulator
import numpy as np
from tabulate import tabulate
from matplotlib import pyplot as plt
from parameter import Parameter
from parameter import Environment_Variable
import parameter as par
from datetime import datetime
from feed_supplier import Supplier
# import pandas as pd
import openpyxl
from openpyxl.styles import Font

variable = 'Speed of build'

Start = True

wb = openpyxl.load_workbook('Results from simulations.xlsx')
print(type(wb))
sheet = wb.create_sheet(title='Sens Start_liq')

values = [0.2, 0.3, 0.5, 1.0]


def cell_write(sheet, coordinates, val, title=False, width='s'):
    assert isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet)
    assert type(coordinates) == tuple and len(coordinates) == 2, ('wrong tuple,', coordinates)
    assert type(coordinates[0]) == int and coordinates[0] > 0
    assert type(coordinates[1]) == int and 27 > coordinates[1] > 0
    assert width == 'w' or width == 'ww' or width == 's', 'width input incorrect. Must be "w" or "ww"'
    row = coordinates[0]
    col = coordinates[1]

    col_letter = chr(col + 64)
    excel_coord = col_letter + str(row)

    _cell = sheet.cell(row=row, column=col)

    if not type(val) == str or type(val):
        if type(val) == np.float64:

            if abs(val) > 1000:
                _cell.number_format = '0.0'
            elif abs(val) > 100:
                _cell.number_format = '0.00'
            elif abs(val) > 10:
                _cell.number_format = '0.000'
            else:
                _cell.number_format = '0.0000'

        # _cell.number_format = '0.0000E+00'

    sheet[excel_coord].value = val

    if width == 'w':
        sheet.column_dimensions[col_letter].width = 15
    elif width == 'ww':
        sheet.column_dimensions[col_letter].width = 40

    if title:
        big_font = Font(size=12, bold=True)
        sheet[excel_coord].font = big_font

    return


def sensitivity_analysis(variable, sheet, values):
    assert type(variable) == str
    assert isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet)
    assert type(values) == list

    months = 240 + 48

    cell_write(sheet, (1, 1), 'Sensitivity analysis', title=True, width='w')

    cell_write(sheet, (3, 1), 'Variable:', title=True)
    cell_write(sheet, (3, 2), variable, title=True)

    cell_write(sheet, (4, 1), ' ')
    cell_write(sheet, (5, 1), 'Final Liquidity')
    cell_write(sheet, (6, 1), 'Starting liquidity')
    cell_write(sheet, (7, 1), 'Average net Profit')
    cell_write(sheet, (8, 1), 'Final emissions')
    cell_write(sheet, (9, 1), 'Average emissions')
    cell_write(sheet, (10, 1), 'bio_proportion')

    for index in range(len(values)):
        # line where the value is changed or this simulation
        sim = simulate(months, table=False, plot=False, Excel_p=False,
                       # manufacturer settings
                       capacity_root_coeff=2.0, speed_of_build=values[index], time_to_build=6.0, fossil_process_cost=1,
                       bio_process_cost=1,
                       starting_liquidity=7000, PET_price=10.0, foss_cap_price=5, bio_cap_price=7,
                       # regulator settings
                       notice_period=30, fraction=0.1, start_levy=1.5, ratio_jump=0.5, wait_time=48,
                       compliance_threshold=0.5, decade_jump=0.5,

                       price_elas=0.1, start_feed_price=2.0)

        final_liquidity = sim[0]
        start_liquidity = sim[1]
        average_net_profit = sim[2]
        final_emissions = sim[3]
        average_emissions = sim[4]
        bio_proportion = sim[5]

        cell_write(sheet, (4, 2 + index), values[index])
        for i in range(len(sim)):
            cell_write(sheet, (5 + i, 2 + index), sim[i])

    wb.save('Results from simulations.xlsx')
    return

sensitivity_analysis(variable, sheet, values)
