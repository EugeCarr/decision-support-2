from matplotlib import pyplot as plt
import openpyxl


def search_var_header(x_header):
    assert type(x_header) == str, ('variable header must be a string, not:', type(x_header))
    found = False
    i = 1
    while not found:
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == x_header:
                x_header_ret = sheet.cell(row=i, column=j)
                found = True
                break

        if i == 20 and not found:
            break

        i += 1

    if found:
        # print (x_header_ret.value)
        return x_header_ret
    if not found:
        raise AssertionError('column header', x_header, 'could not be found')
    return


def get_values(header, highest_row):
    assert isinstance(header, openpyxl.cell.cell.Cell), ('header must be type cell not:', type(header))
    assert type(highest_row) == int, ('highest row must be an integer,not:', type(highest_row))
    values = []

    for row in range(header.row + 1, highest_row + 1):
        cell_val = sheet.cell(row=row, column=header.column).value

        assert type(cell_val) != str
        values.append(cell_val)

    return values


wb = openpyxl.load_workbook('Results from simulations.xlsx')

sheet = wb['Reactive regulator']

x_var_name = 'Month'

y_var_name = 'levy_rate'

y2_var_name = 'fossil_capacity'

y3_var_name = 'fossil_capacity'

x_var_header = search_var_header(x_var_name)
x_values = get_values(x_var_header, sheet.max_row)

y_var_header = search_var_header(y_var_name)
y_values = get_values(y_var_header, sheet.max_row)

y2_var_header = search_var_header(y2_var_name)
y2_values = get_values(y2_var_header, sheet.max_row)

y3_var_header = search_var_header(y3_var_name)
y3_values = get_values(y3_var_header, sheet.max_row)

same_axes = True

if not same_axes:

    fig, ax1 = plt.subplots()
    ax1.plot(x_values, y_values, label='Levy rate')
    ax1.set_ylabel('Levy rate')
    # ax1.plot(x_values, y_values, label='Net profit')
    # ax1.set_ylabel('Net profit')
    ax1.set_xlabel(x_var_name)

    # ax1.set_ylabel('Liquidity')
    # h1, l1 = ax1.get_legend_handles_labels()

    ax2 = ax1.twinx()  # instantiate a second axes that shares the same x-axis

    # ax2.set_ylabel('Proportion bio')  # we already handled the x-label with ax1
    # h2, l2 = ax2.get_legend_handles_labels()
    ax2.plot(x_values, y2_values, 'm--', label='Fossil fuel based production')
    ax2.plot(x_values, y3_values, 'r--', label='Fossil fuel based capacity')
    ax2.set_ylabel('Fossil fuel based production')

    # ax2.tick_params(axis='y')
    #
    # ax2.set_xlabel(x_var_name)

    # plt.legend(handles=[p1, p2], loc="best")
    # plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', borderaxespad=0.)
    ax2.legend(fontsize=8, loc=(0.50, -0.15), frameon=False)
    # ax1.legend(loc=0, frameon=False)
    ax1.legend(loc=(0.00, -0.15), fontsize=8, frameon=False)

    plt.tick_params(labelsize=8)
    ax1.tick_params(labelsize=8)



    fig.tight_layout()
    plt.show()

else:
    fig, ax1 = plt.subplots()
    # ax1.plot(x_values, y_values, label='Fossil fuel based capacity')
    ax1.plot(x_values, y_values,  'b--', label='Levy rate')
    ax1.set_ylabel('Levy rate')
    ax1.set_xlabel(x_var_name)

    # ax1.legend(loc=(0.00, 0.95), fontsize=8, frameon=False)

    plt.tick_params(labelsize=8)
    ax1.tick_params(labelsize=8)



    fig.tight_layout()
    plt.show()
