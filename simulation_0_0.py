"""This file defines simulation parameters for the first build of the model"""
import agent as ag
from regulator import Regulator
import numpy as np
from tabulate import tabulate
from matplotlib import pyplot as plt
from parameter import Parameter
from parameter import Environment_Variable
import parameter as par
from datetime import datetime
from feed_supplier import Supplier
# import pandas as pd
import openpyxl
from openpyxl.styles import Font


def simulate(months, table=False, plot=True, Excel_p=False):
    # def simulate(months, table=False, plot=False):
    # create agents and specify their parameters
    month = int(0)
    initial_production_volume = np.float64(1000)

    # the dictionary of environment variables (see parameter.py) to pass to the Environment object
    env_variables = {
        'pet_price': Environment_Variable(par.pet_price, months, init=np.float64(10.0)),
        'fossil_feedstock_price': Environment_Variable(par.fossil_feedstock_price, months, init=np.float64(3)),
        'bio_feedstock_price': Environment_Variable(par.bio_feedstock_price, months, init=np.float64(2)),
        'levy_rate': Environment_Variable(par.levy_rate, months, init=np.float64(0.0)),
        'demand': Environment_Variable(par.demand, months, init=np.float64(1000))
    }

    env_keys = list(env_variables.keys())

    env_aggregates = {
        'fossil_feedstock_consumption': Environment_Variable(par.blank, months),
        'bio_feedstock_consumption': Environment_Variable(par.blank, months),
        'emissions': Environment_Variable(par.blank, months)
    }

    env_aggregates_keys = list(env_aggregates.keys())

    environment = ag.Environment(env_variables, env_aggregates)

    # dictionary of all variables in the order in which they should be computed
    # parameters from the environment that need to be projected by the agent use par.blank for the fun argument
    # similarly for parameters calculated by the agent but which are not projected
    manufacturer1_parameters = {
        'demand': Parameter(par.blank, par.demand_projection, months),
        'bio_capacity_max': Parameter(par.bio_capacity_max, par.bio_capacity_max_projection, months),
        'fossil_capacity_max': Parameter(par.fossil_capacity_max, par.fossil_capacity_max_projection, months),

        'fossil_capacity': Parameter(par.fossil_capacity_alt2, par.fossil_capacity_projection_alt2, months,
                                     init=initial_production_volume),
        'bio_capacity': Parameter(par.bio_capacity_alt2, par.bio_capacity_projection_alt2, months),
        'expansion_cost': Parameter(par.expansion_cost, par.expansion_cost_projection, months),

        'fossil_production': Parameter(par.fossil_production, par.fossil_production_projection, months,
                                       init=np.float64(1000)),
        'bio_production': Parameter(par.bio_production, par.bio_production_projection, months),
        'total_production': Parameter(par.total_production, par.total_production_projection, months,
                                      init=np.float64(1000)),

        'fossil_feedstock_consumption': Parameter(par.fossil_feedstock_consumption,
                                                  par.fossil_feedstock_consumption_projection, months),
        'bio_feedstock_consumption': Parameter(par.bio_feedstock_consumption,
                                               par.bio_feedstock_consumption_projection, months),
        'unit_sale_price': Parameter(par.blank, par.unit_sale_price_projection, months),
        'fossil_feedstock_price': Parameter(par.blank, par.fossil_feedstock_price_projection, months,
                                            init=np.float64(3)),
        'bio_feedstock_price': Parameter(par.blank, par.bio_feedstock_price_projection, months),


        # 'production_volume': Parameter(par.production_volume, par.production_volume_projection, months,
        #                                init=initial_production_volume),

        'fossil_process_cost': Parameter(par.fossil_process_cost, par.fossil_process_cost_projection, months,
                                         init=np.float64(1)),
        'bio_process_cost': Parameter(par.bio_process_cost, par.bio_process_cost_projection, months,
                                      init=np.float64(0.5)),
        'emissions': Parameter(par.emissions, par.emissions_projection, months),
        'levy_rate': Parameter(par.blank, par.levy_rate_projection, months, init=np.float64(0.2)),
        'levies_payable': Parameter(par.levies_payable, par.levies_payable_projection, months),

        'gross_profit': Parameter(par.gross_profit, par.gross_profit_projection, months),
        'tax_payable': Parameter(par.tax_payable, par.tax_payable_projection, months),
        'net_profit': Parameter(par.net_profit, par.net_profit_projection, months),

        'profitability': Parameter(par.profitability, par.profitability_projection, months),
        'liquidity': Parameter(par.liquidity, par.liquidity_projection, months, init=np.float64(5000)),
        'profit_margin': Parameter(par.profit_margin, par.profit_margin_projection, months)
    }

    manufacturer1 = ag.Manufacturer('PET Manufacturer 1', months, environment, manufacturer1_parameters,
                                    capacity_root_coefficient=1.25, time_to_build=6.0)

    regulator = Regulator(name='Regulator', sim_time=months, env=environment, tax_rate=0.19, notice_period=12,
                          fraction=0.5, start_levy=1.0, ratio_jump=0.5, wait_time=120, compliance_threshold=0.5, decade_jump=0.2)

    supplier = Supplier('supplier', months, environment, 2.0, elasticity=0.5)

    manufacturers = [
        manufacturer1
    ]

    agents = manufacturers + [regulator, supplier]

    sim_start = datetime.now()
    # Run simulation for defined number of months
    while month < months:
        for key in env_keys:
            if month != 0:
                environment.parameter[key].update(environment)

            environment.parameter[key].record(month)

        # advance time counter in environment
        environment.month = month
        # advance time counter in each agent
        for agent in agents:
            agent.month = month

        # execute monthly routines on manufacturers
        for agent in manufacturers:
            agent.time_step_alt()



        environment.reset_aggregates()
        for key in env_aggregates_keys:
            for manufacturer in manufacturers:
                try:
                    environment.aggregate[key].value += manufacturer.parameter[key].value
                except KeyError:
                    pass

        environment.aggregate['bio_feedstock_consumption'].value += 500

        supplier.iterate_supplier()
        regulator.iterate_regulator()

        for key in env_aggregates_keys:
            environment.aggregate[key].record(month)



        # if the regulator rate has just changed then update it in the environment
        if environment.parameter['levy_rate'].value != regulator.levy_rate:
            environment.parameter['levy_rate'].value = regulator.levy_rate
            environment.time_to_levy_change = regulator.time_to_change()
            environment.levy_rate_changing = False

        # if a change in the levy rate is approaching, add this information to the environment
        if regulator.change_check():
            environment.levy_rate_changing = True
            environment.time_to_levy_change = regulator.time_to_change()
            environment.future_levy_rate = regulator.future_levy_rate
        else:
            pass

        month += 1

    sim_end = datetime.now()
    elapsed = sim_end - sim_start
    print('\n Simulation elapsed time:', elapsed)
    print('\n ============ \n FINAL STATE \n ============',
          '\n Regulation level:', regulator.level,
          '\n Levy rate:', environment.parameter['levy_rate'].value,
          '\n Bio capacity 1:', manufacturer1.parameter['bio_capacity'].value)
    print(' Target:', manufacturer1.bio_capacity_target)

    # data output & analysis
    t = np.arange(0, months, 1)

    if table:
        table = []
        for i in range(0, months):
            table.append([t[i],
                          environment.parameter['levy_rate'].history[i]])

        headers = ['Month', 'levy_rate']
        print(tabulate(table, headers))

    if plot:
        graph(environment.parameter['levy_rate'])
        graph(manufacturer1.parameter['fossil_capacity'])
        # graph(environment.parameter['fossil_feedstock_price'])
        graph(manufacturer1.parameter['fossil_production'])
        graph(manufacturer1.parameter['bio_capacity'])
        graph(manufacturer1.parameter['bio_production'])
        graph(manufacturer1.parameter['net_profit'])

        # graph(environment.aggregate['emissions'])
        # graph(environment.parameter['demand'])

    if Excel_p:
        # bio_proportion_list = np.divide(manufacturer1.parameter['bio_production'].history,
        #                                 (manufacturer1.parameter['bio_production'].history + manufacturer1.parameter[
        #                                     'fossil_production'].history))

        # when the changes from rewrite_optimisation are merged in, a new parameter needs to be made for bio_proportion
        wb = openpyxl.load_workbook('Results from simulations.xlsx')
        # print(type(wb))
        sheet = wb.create_sheet(title='smoother fast build')
        # print(sheet.title)
        date_time = datetime.now()

        cell_write(sheet, (1, 1), 'Simulation of decision support tool', title=True, width='w')

        cell_write(sheet, (3, 1), 'General information', title=True)
        cell_write(sheet, (4, 1), 'Date & time:', title=True)
        cell_write(sheet, (4, 2), date_time, title=False, width='w')
        cell_write(sheet, (5, 1), 'No. of Manufacturers', title=True)
        cell_write(sheet, (5, 2), len(manufacturers))
        # cell_write(sheet, (6, 1), 'Run time')
        # cell_write(sheet, (6, 2), elapsed)

        cell_write(sheet, (3, 4), 'Regulator Settings', title=True)
        cell_write(sheet, (4, 4), 'Tax rate', title=True)
        cell_write(sheet, (4, 5), regulator.tax_rate)
        cell_write(sheet, (5, 4), 'Emission fraction width', title=True)
        cell_write(sheet, (5, 5), regulator.fraction)
        cell_write(sheet, (6, 4), 'Emission ratio jump', title=True)
        cell_write(sheet, (6, 5), regulator.ratio_jump)
        cell_write(sheet, (7, 4), 'Initial levy rate', title=True)
        cell_write(sheet, (7, 5), regulator.start_levy)
        cell_write(sheet, (8, 4), 'Decade Change', title=True)
        cell_write(sheet, (8, 5), regulator.dec_jump)

        cell_write(sheet, (3, 6), 'Supplier Settings', title=True, width='w')
        cell_write(sheet, (4, 6), 'Initial price', title=True)
        cell_write(sheet, (4, 7), supplier.start_price)
        cell_write(sheet, (5, 6), 'Price elasticity', title=True)
        cell_write(sheet, (5, 7), supplier.price_elasticity)

        cell_write(sheet, (3, 8), 'Manufacturer 1 settings', title=True, width='w')
        cell_write(sheet, (4, 8), 'Starting liquidity', title=True)
        cell_write(sheet, (4, 9), manufacturer1.parameter['liquidity'].history[0])
        cell_write(sheet, (5, 8), 'Bio process cost', title=True)
        cell_write(sheet, (5, 9), manufacturer1.parameter['bio_process_cost'].history[0])
        cell_write(sheet, (7, 8), 'Fossil process cost', title=True)
        cell_write(sheet, (7, 9), manufacturer1.parameter['fossil_process_cost'].history[0])
        cell_write(sheet, (6, 8), 'Bio feedstock price', title=True)
        cell_write(sheet, (6, 9), environment.parameter['bio_feedstock_price'].history[0])
        cell_write(sheet, (8, 8), 'Fossil feedstock price', title=True)
        cell_write(sheet, (8, 9), environment.parameter['fossil_feedstock_price'].history[0])
        cell_write(sheet, (9, 8), 'Starting Production', title=True)
        cell_write(sheet, (9, 9), manufacturer1.parameter['total_production'].history[0])
        cell_write(sheet, (10, 8), 'Initial PET price', title=True)
        cell_write(sheet, (10, 9), environment.parameter['pet_price'].history[0])
        cell_write(sheet, (4, 10), 'Maintenance cost', title=True)
        cell_write(sheet, (4, 11), manufacturer1.capacity_maintenance_cost)
        cell_write(sheet, (5, 10), 'Bio capacity cost', title=True, width='w')
        cell_write(sheet, (5, 11), manufacturer1.bio_capacity_cost)
        cell_write(sheet, (6, 10), 'Fossil capacity cost', title=True, width='w')
        cell_write(sheet, (6, 11), manufacturer1.fossil_capacity_cost)
        cell_write(sheet, (8, 10), 'Build speed', title=True, width='w')
        cell_write(sheet, (8, 11), manufacturer1.build_speed)
        cell_write(sheet, (7, 10), 'Root capacity coefficient', title=True, width='ww')
        cell_write(sheet, (7, 11), manufacturer1.capacity_root_coefficient)

        variables = [
            ('m1', 'fossil_production'),
            ('m1', 'bio_production'),
            ('m1', 'liquidity'),
            ('m1', 'profitability'),
            ('E', 'levy_rate'),
            ('EA', 'emissions'),
            ('E', 'bio_feedstock_price'),
            ('m1', 'bio_capacity'),
            ('m1', 'fossil_capacity'),
            ('E', 'demand')
        ]
        # var2 : ('m1', 'bio_production'),
        # var3 : ('m1', 'fossil_production')

        cell_write(sheet, (12, 1), 'Month', title=True)

        for i in range(1, months + 1):
            cell_write(sheet, (12 + i, 1), i)
            _cell = sheet.cell(row=(13 + i), column=1)
            _cell.number_format = '0'

        history = []
        for i in range(len(variables)):
            variable = variables[i]
            if variable[0] == 'm1':
                history = list(manufacturer1.parameter[variable[1]].history)
            elif variable[0] == 'E':
                history = list(environment.parameter[variable[1]].history)
            elif variable[0] == 'EA':
                history = list(environment.aggregate[variable[1]].history)
            else:
                print("invalid variable name,", variable[0])

            data_column = 2 * (i + 1)

            cell_write(sheet, (12, data_column), variable[1], title=True, width='w')
            for j in range(len(history)):
                cell_write(sheet, (13 + j, data_column), history[j])

        wb.save('Results from simulations.xlsx')
    return


def graph(parameter):
    assert isinstance(parameter, Parameter) or isinstance(parameter, Environment_Variable)
    y = parameter.history
    t = np.arange(0, len(y), 1)
    fig, ax1 = plt.subplots()
    ax1.plot(t, y)

    ax1.set_xlabel('Month')
    # ax1.set_ylabel('Price of bio feedstock')
    # ax1.set_ylabel('Capacity bio')
    ax1.set_ylabel('Capacity')
    # ax1.set_ylabel('Proportion bio-PET')
    # ax1.set_ylabel('Levy rate')
    # ax1.set_ylabel('Demand')
    # ax1.set_ylabel('Emissions')


    fig.tight_layout()
    plt.show()
    return


def graph2(parameter1, parameter2):
    assert isinstance(parameter1, Parameter) or isinstance(parameter1, Environment_Variable)
    assert isinstance(parameter2, Parameter) or isinstance(parameter2, Environment_Variable)
    y1 = parameter1.history
    t1 = np.arange(0, len(y1), 1)
    y2 = parameter2.history
    t2 = np.arange(0, len(y2), 1)

    fig, ax1 = plt.subplots()

    color = 'tab:red'
    ax1.set_xlabel('time (months)')
    ax1.set_ylabel('', color=color)
    ax1.plot(t1, y1, color=color)
    ax1.tick_params(axis='y', labelcolor=color)

    ax2 = ax1.twinx()  # instantiate a second axes that shares the same x-axis

    color = 'tab:blue'
    ax2.set_ylabel('', color=color)  # we already handled the x-label with ax1
    ax2.plot(t2, y2, color=color)
    ax2.tick_params(axis='y', labelcolor=color)

    fig.tight_layout()  # otherwise the right y-label is slightly clipped
    plt.show()
    return


def cell_write(sheet, coordinates, val, title=False, width='s'):
    assert isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet)
    assert type(coordinates) == tuple and len(coordinates) == 2, ('wrong tuple,', coordinates)
    assert type(coordinates[0]) == int and coordinates[0] > 0
    assert type(coordinates[1]) == int and 27 > coordinates[1] > 0
    assert width == 'w' or width == 'ww' or width == 's', 'width input incorrect. Must be "w" or "ww"'
    row = coordinates[0]
    col = coordinates[1]

    col_letter = chr(col + 64)
    excel_coord = col_letter + str(row)

    _cell = sheet.cell(row=row, column=col)

    if not type(val) == str or type(val):
        if type(val) == np.float64:

            if abs(val) > 1000:
                _cell.number_format = '0.0'
            elif abs(val) > 100:
                _cell.number_format = '0.00'
            elif abs(val) > 10:
                _cell.number_format = '0.000'
            else:
                _cell.number_format = '0.0000'

        # _cell.number_format = '0.0000E+00'

    sheet[excel_coord].value = val

    if width == 'w':
        sheet.column_dimensions[col_letter].width = 15
    elif width == 'ww':
        sheet.column_dimensions[col_letter].width = 40

    if title:
        big_font = Font(size=12, bold=True)
        sheet[excel_coord].font = big_font

    return
